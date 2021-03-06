import os
import datetime
import openpyxl



def sheet_edit(workbook,date):  #workbook, (str)DATE
    if date not in workbook.sheetnames:
        workbook.copy_worksheet(workbook.worksheets[0])
        workbook.worksheets[-1].title = date
    return workbook[date]



def edit(row, workbook):
    #insert row check
    start_time = datetime.datetime.strptime(row[2],"%Y/%m/%d %H:%M")
    st_t = datetime.datetime.strftime(start_time,"%H:%M")
    end_time = datetime.datetime(1,1,1)
    en_t = ''
    if row[0] in ["1","2","3","4"]:
        end_time = datetime.datetime.strptime(row[4],"%Y/%m/%d %H:%M")
        en_t = datetime.datetime.strftime(end_time,"%H:%M")


    #worksheet select
    worksheet = sheet_edit(workbook,datetime.datetime.strftime(start_time,"%Y-%m-%d"))


    #collupse check
    row_write = 1
    while worksheet.cell(row=row_write,column=1).value==None or worksheet.cell(row=row_write,column=1).value < datetime.datetime.strftime(start_time,"%H:%M"):
        row_write += 1
        if row_write > worksheet.max_row:
            break


    #duplicate check
    row_duplicate_check = row_write
    while row_duplicate_check<=worksheet.max_row:
        if worksheet.cell(row=row_duplicate_check,column=1).value!=None and worksheet.cell(row=row_duplicate_check,column=1).value < datetime.datetime.strftime(end_time,"%H:%M"):
            print("時間重複")
            return
        row_duplicate_check += 1


    #insert rows
    rows_numbers = [0,4,3,3,3]
    worksheet.insert_rows(row_write,rows_numbers[int(row[0])])

    #alignments settings
    for i in range(rows_numbers[int(row[0])]):
        worksheet.cell(row=row_write+i,column=1).alignment = openpyxl.styles.Alignment(horizontal='right')
        worksheet.cell(row=row_write+i,column=3).alignment = openpyxl.styles.Alignment(horizontal='center')


    #write
    if row[0] =="1":
        worksheet.cell(row=row_write,column=4).value = row[1]       #発駅
        worksheet.cell(row=row_write,column=1).value = st_t         #発時刻
        worksheet.cell(row=row_write+3,column=4).value = row[3]     #着駅
        worksheet.cell(row=row_write+3,column=1).value = en_t       #着時刻
        worksheet.cell(row=row_write+1,column=5).value = row[5]     #路線
        worksheet.cell(row=row_write+1,column=6).value = row[6]     #列車番号
        worksheet.cell(row=row_write+2,column=5).value = row[7]     #列車種別
        worksheet.cell(row=row_write+2,column=6).value = row[8]     #行先
        worksheet.cell(row=row_write,column=3).value = row[9]       #発番線
        worksheet.cell(row=row_write+3,column=3).value = row[10]    #着番線
